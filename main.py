import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font


def create_progression_sheet(wb, sheet_name, setting_cell):
    """
    Create a progression sheet in the workbook 'wb' with the given sheet name.
    The progression formulas refer to the max PR stored in the settings sheet at setting_cell.
    """
    ws = wb.create_sheet(title=sheet_name)
    headers = ["Week", "Day", "Session Type", "Sets", "Factor", "Reps per Set", "Volume"]
    ws.append(headers)

    # Set header style (bold)
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font

    # Define progression factors for each week (Blocks: Weeks 1-4, 5-8, 9-12)
    factors = {
        1: 0.55, 2: 0.65, 3: 0.75, 4: 0.70,
        5: 0.60, 6: 0.70, 7: 0.80, 8: 0.75,
        9: 0.65, 10: 0.75, 11: 0.85, 12: 1.00
    }

    # Define default session types for weeks 1-11:
    # Format: (Session Name, Sets, Adjustment to target reps)
    default_sessions = [
        ("Standard", 4, 0),
        ("Volume Low", 6, -0.2),
        ("Standard", 4, 0),
        ("Volume High", 5, -0.1)
    ]
    # Adjusted to work between 4-7 reps with higher set counts.

    current_row = 2  # start after header
    block_weekly_rows = []  # to store weekly summary row numbers for current block
    block_number = 1

    for week in range(1, 13):
        week_start_row = current_row

        # For week 12, override the session structure:
        if week == 12:
            sessions_for_week = [
                ("Standard", 4, 0),
                ("Light", 4, -0.2),
                ("Very Light", 4, -0.3),
                ("Test Day", 1, 0)  # Test Day: no formula, just a prompt
            ]
        else:
            sessions_for_week = default_sessions

        # Write the four training day rows for the week
        day_index = 1
        for session in sessions_for_week:
            session_name, sets_val, adjustment = session
            ws.append([week, f"Day {day_index}", session_name, sets_val, factors[week], None, None])
            cell_rep = ws.cell(row=current_row, column=6)
            cell_vol = ws.cell(row=current_row, column=7)
            if session_name == "Test Day":
                cell_rep.value = "Test"
                cell_vol.value = ""
            else:
                # Create formula for "Reps per Set" using the factor in column E.
                # The formula calculates:
                formula = f"=INT('Settings'!{setting_cell}*E{current_row}"
                if adjustment > 0:
                    formula += f" + INT('Settings'!{setting_cell}*{adjustment})"
                elif adjustment < 0:
                    formula += f" - INT('Settings'!{setting_cell}*{abs(adjustment)})"
                formula += ")"
                cell_rep.value = formula
                # Volume = Sets * Reps per Set
                cell_vol.value = f"=D{current_row}*F{current_row}"
            current_row += 1
            day_index += 1

        # Add weekly summary row (total volume for the week)
        week_end_row = current_row - 1
        ws.append([f"Week {week} Total Volume", "", "", "", "", "", f"=SUM(G{week_start_row}:G{week_end_row})"])
        weekly_summary_row = current_row
        block_weekly_rows.append(weekly_summary_row)
        current_row += 1  # move to next row

        # If end of a block (every 4 weeks) or end of plan, add block summary and separator
        if week % 4 == 0 or week == 12:
            # Create block summary row summing the weekly totals in this block.
            sum_refs = ",".join([f"G{r}" for r in block_weekly_rows])
            ws.append([f"Block {block_number} Total Volume", "", "", "", "", "", f"=SUM({sum_refs})"])
            block_summary_row = current_row
            # Apply a light fill for visibility on the block summary row
            block_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            for col in range(1, 8):
                ws.cell(row=block_summary_row, column=col).fill = block_fill
            current_row += 1  # next row
            # Add an extra blank row for separation
            ws.append([""] * 7)
            current_row += 1
            block_weekly_rows = []  # reset for the next block
            block_number += 1

    # Hide the Factor column (Column E)
    ws.column_dimensions['E'].hidden = True

    # Adjust column widths for readability
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                length = len(str(cell.value))
                if length > max_length:
                    max_length = length
        ws.column_dimensions[col_letter].width = max_length + 1


def main():
    wb = Workbook()

    # Create Settings sheet with pull-up and push-up settings
    ws_settings = wb.active
    ws_settings.title = "Settings"

    # Pull-Up Settings
    ws_settings["A1"] = "Pull-Up Progression Settings"
    ws_settings["A2"] = "Enter your Max Pull-Up PR in cell C2"
    ws_settings["B2"] = "Max PR"
    ws_settings["C2"] = 10  # default pull-up max

    # Push-Up Settings
    ws_settings["A4"] = "Push-Up Progression Settings"
    ws_settings["A5"] = "Enter your Max Push-Up PR in cell C4"
    ws_settings["B5"] = "Max PR"
    ws_settings["C4"] = 25  # default push-up max

    # Create progression sheets for push-ups and pull-ups.
    # For pull-ups, use the setting in C2; for push-ups, use the setting in C4.
    create_progression_sheet(wb, "PullUp Progression", "$C$2")
    create_progression_sheet(wb, "PushUp Progression", "$C$4")

    # Save the workbook
    wb.save("Progression.xlsx")


if __name__ == "__main__":
    main()
